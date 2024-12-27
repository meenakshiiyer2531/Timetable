import openpyxl
from pprint import pprint
from timetable_class import Timetable
from openpyxl.styles import Font,Alignment,PatternFill

class Teacher:
    """Represents a teacher with details like ID, name, subjects, preferences, and timetable."""
    def __init__(self, teacher_id, name, subjects, preferred_sections, is_assistant=False, assigned_section=None):
        """
        Initializes a Teacher object with the provided details.

        Args:
            teacher_id (str): Unique identifier for the teacher.
            name (str): Name of the teacher.
            subjects (list): List of subject codes the teacher can teach.
            preferred_sections (list): List of preferred section IDs.
            is_assistant (bool, optional): Flag indicating if the teacher is an assistant. Defaults to False.
        """
        self.teacher_id = teacher_id  # Unique identifier for the teacher
        self.name = name  # Name of the teacher
        self.subjects = subjects  # List of subjects the teacher can teach
        self.preferred_sections = preferred_sections  # List of preferred sections
        self.is_assistant = is_assistant  # Flag indicating if the teacher is an assistant
        self.assigned_section = assigned_section  # Section ID if the teacher is a class animator

    def print_data(self):
        print(f" {self.teacher_id} " , end="")
        print(f" {self.name} " , end="")
        print(f" {self.subjects} " , end="")
        print(f" {self.preferred_sections} " , end="")
        print(f" {self.is_assistant} " , end="")
        print(f" {self.assigned_section} " , end="") 
        print("\n")

        # print(f"Type - teacher_id - {type(self.teacher_id) } " , end="")
        # print(f"Type - teacher_name - {type(self.name) } " , end="")
        # print(f"Type - subjects - {type(self.subjects) } " , end="")
        # print(f"Type - preferred_sections - {type(self.preferred_sections) } " , end="")
        # print(f"Type - is_assistant - {type(self.is_assistant) } " , end="")
        # print(f"Type - assigned_class - {type(self.assigned_section) } " , end="") 
        # print("\n")


class Subject:
    """Represents a subject with details like code, name, repetitions, and type."""
    def __init__(self, subject_code, name, weekly_repetitions, is_practical, is_extracurricular,year,specialization):
        """
        Initializes a Subject object with the provided details.

        Args:
            subject_code (str): Unique code for the subject.
            name (str): Name of the subject.
            weekly_repetitions (int): Number of times the subject should be taught per week.
            is_practical (bool): Flag indicating if the subject is practical.
            is_extracurricular (bool): Flag indicating if the subject is extracurricular.
        """
        self.subject_code = subject_code  # Unique code for the subject
        self.name = name  # Name of the subject
        self.weekly_repetitions = weekly_repetitions  # Number of times the subject should be taught per week
        self.is_practical = is_practical  # Flag indicating if the subject is practical
        self.is_extracurricular = is_extracurricular  # Flag indicating if the subject is (eg. CA) extracurricular
        self.year = year
        self.specialization = specialization

    def print_data(self):
        print(f" {self.subject_code} " , end="")
        print(f" {self.name} " , end="")
        print(f" {self.weekly_repetitions} " , end="")
        print(f" {self.is_practical} " , end="")
        print(f" {self.is_extracurricular} " , end="")
        print(f" {self.year} " , end="")
        print(f" {self.specialization} " , end="")
        print("\n")
        
        # print(f"Type - subject_code - {type(self.subject_code)} " , end="")
        # print(f"Type - name - {type(self.name)} " , end="")
        # print(f"Type - weekly_repetitions - {type(self.weekly_repetitions)} " , end="")
        # print(f"Type - is_practical - {type(self.is_practical)} " , end="")
        # print(f"Type - is_extracurricular - {type(self.is_extracurricular)} " , end="")
        # print(f"Type - year - {type(self.year)} " , end="")
        # print(f"Type - specialization - {type(self.specialization)} " , end="")
        # print("\n")


class Section:
    """Represents a section with details like ID, name, students, batch, and timetable."""
    def __init__(self, section_id, name, students, batch,year,specialization,preferred_room = None):
        """
        Initializes a Section object with the provided details.

        Args:
            section_id (str): Unique identifier for the section.
            name (str): Name of the section.
            students (list): List of student names in the section.
            batch (int): Batch of the section (0 for morning, 1 for evening).
        """
        self.section_id = section_id  # Unique identifier for the section
        self.name = name  # Name of the section
        self.students = students  # List of students in the section
        self.batch = batch  # Batch of the section (0 for morning, 1 for evening)
        self.year = year
        self.specialization = specialization
        self.preferred_room = preferred_room # Preferred room if any


    def print_data(self):
        print(f" {self.section_id} " , end="")
        print(f" {self.name} " , end="")
        print(f" {self.students} " , end="")
        print(f" {self.batch} " , end="")
        print(f" {self.year} " , end="")
        print(f" {self.specialization} " , end="")
        print(f" {self.preferred_room} " , end="")
        print("\n")

        # print(f"Type - section_id - {type(self.section_id)} " , end="")
        # print(f"Type - name - {type(self.name)} " , end="")
        # print(f"Type - students - {type(self.students)} " , end="")
        # print(f"Type - batch - {type(self.batch)} " , end="")
        # print(f"Type - year - {type(self.year)} " , end="")
        # print(f"Type - specialisation - {type(self.specialization)} " , end="")
        # print(f"Type - preferred_room - {type(self.preferred_room)} " , end="")
        # print("\n")
        

class Room:
    def __init__(self, room_number, is_lab=False ,capacity=None):
        self.room_number = room_number
        self.is_lab = is_lab
        self.capacity = capacity

    
    def print_data(self):
        print(f"room_number - {self.room_number} " , end="")
        print(f"is_lab - {self.is_lab} " , end="")
        print(f"capacity - {self.capacity} " , end="")
        print("\n")
        
        # print(f"Type - room_number - {type(self.room_number)} " , end="")
        # print(f"Type - is_lab - {type(self.is_lab)} " , end="")
        # print(f"Type - capacity - {type(self.capacity)} " , end="")
        # print("\n")
        

class read_data: 
    def __init__(self, file_path_add):
        self.file_path=file_path_add

    def read_excel_data_teacher(self):
        workbook = openpyxl.load_workbook(self.file_path)
        teachers_sheet = workbook['Teachers']

        teachers = []

        row=3
        while teachers_sheet[f'A{row}'].value is not None:
            teacher_id=str(teachers_sheet[f'A{row}'].value)
            teacher_name=str(teachers_sheet[f'B{row}'].value)
            subjects =teachers_sheet[f'C{row}'].value
            preferred_sections=teachers_sheet[f'D{row}'].value
            is_assistant =teachers_sheet[f'E{row}'].value
            assigned_class=str(teachers_sheet[f'F{row}'].value)
            row+=1

            # Convert strings to lists
            subjects_lst = [x1.strip() for x1 in subjects.split(",")]  # Split comma-separated string
            preferred_sections_lst = [x1.strip() for x1 in preferred_sections.split(",")]  # Split comma-separated string
            is_assistant_bool = str(is_assistant).lower() == "true"  # Convert to boolean

            teacher = Teacher(teacher_id,teacher_name, subjects_lst,preferred_sections_lst,is_assistant_bool,assigned_class)
            # teacher.print_data()
            teachers.append(teacher)

        return teachers

    def read_excel_data_section(self):
        workbook = openpyxl.load_workbook(self.file_path)
        sections_sheet = workbook['Sections']

        sections = []
        # Data type needed for Sections - String,String,list,int,int,string,string
        row=3
        while sections_sheet[f'A{row}'].value is not None:
            section_id=str(sections_sheet[f'A{row}'].value)
            section_name=str(sections_sheet[f'B{row}'].value)
            student_lst =[sections_sheet[f'C{row}'].value]
            batch=int(sections_sheet[f'D{row}'].value)
            year =int(sections_sheet[f'E{row}'].value)
            specialisation=str(sections_sheet[f'F{row}'].value)
            preferred_room=str(sections_sheet[f'G{row}'].value)
            row+=1

            # Data type needed - String,String,list,int,int,string,string
            section = Section(section_id,section_name, student_lst,batch,year,specialisation,preferred_room)
            # section.print_data()
            sections.append(section)

        return sections

    def read_excel_data_subject(self):
        workbook = openpyxl.load_workbook(self.file_path)
        subjects_sheet = workbook['Subjects']

        subjects = []

        row=3
        # Data type needed for subjects - String,String,int,bool,bool,int,list
        while subjects_sheet[f'A{row}'].value is not None:
            subject_code=str(subjects_sheet[f'A{row}'].value)
            subject_name=str(subjects_sheet[f'B{row}'].value)
            weekely_rep =int(subjects_sheet[f'C{row}'].value)
            is_practical= str(subjects_sheet[f'D{row}'].value).lower() == "true"  # Convert to boolean
            is_ec = str(subjects_sheet[f'E{row}'].value).lower() == "true"  # Convert to boolean
            year=int(subjects_sheet[f'F{row}'].value)
            specialization=[x1.strip() for x1 in ((subjects_sheet[f'G{row}'].value).split(','))]
            row+=1

            # Data type needed - String,String,int,bool,bool,int,list

            subject = Subject(subject_code,subject_name, weekely_rep,is_practical,is_ec,year,specialization)
            # subject.print_data()
            subjects.append(subject)

        return subjects

    def read_excel_data_room(self):
        workbook = openpyxl.load_workbook(self.file_path)
        rooms_sheet = workbook['Rooms']

        rooms = []
        # Data type needed for Rooms - String,Bool,int
        row=3
        while rooms_sheet[f'A{row}'].value is not None:
            room_no=str(rooms_sheet[f'A{row}'].value)
            is_lab=str(rooms_sheet[f'B{row}'].value).lower() == "true"  # Convert to boolean
            capacity=int(rooms_sheet[f'C{row}'].value)
            row+=1

            room = Room(room_no,is_lab, capacity)
            # room.print_data()
            rooms.append(room)

        return rooms

    def read_all_data(self):

        teachers = self.read_excel_data_teacher()
        sections = self.read_excel_data_section()
        subjects = self.read_excel_data_subject()
        rooms = self.read_excel_data_room()

        return teachers,subjects,sections,rooms
    

class excel_output:
    def __init__(self,timetable,file_path):
        self.sections_timetable = timetable.section_timetable
        self.teachers_timetable = timetable.teacher_timetable
        self.rooms_timetable = timetable.room_timetable
        self.file_path = file_path

    
    def change_font(self,worksheet,cell_path,is_section=False,isnormalvalue=False):

        if is_section:
            a1 = worksheet[cell_path]
            a1.font = Font(name='Arial Black',bold=True,size=15) # the change only affects A1
            a1.alignment = Alignment(horizontal="center",vertical="center")
            a1.fill = PatternFill(fgColor="EE7171",fill_type = "solid")
            return
        if isnormalvalue:
            c1 = worksheet[cell_path]
            c1.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            return

        b1 = worksheet[cell_path]
        b1.font = Font(name='Arial Black',bold=True,size=10) # the change only affects b1
        b1.alignment = Alignment(horizontal="center",vertical="center")
        b1.fill = PatternFill(fgColor="AEEC7E",fill_type = "solid")
    
    def clear_sheet(self,sheet_name):

        wb = openpyxl.load_workbook(self.file_path)
        sheetname = sheet_name

        # index of [sheet_name] sheet
        idx = wb.sheetnames.index(sheetname)

        # for new versions, tested with 3.0.3
        ws = wb[sheetname]
        wb.remove(ws)

        # create an empty sheet [sheetname] using old index
        wb.create_sheet(sheetname, idx)

        wb.save(self.file_path)

    def print_output_excel_section(self):
        
        self.clear_sheet("Section Timetable")
        workbook = openpyxl.load_workbook(self.file_path)
        section_sheet = workbook['Section Timetable']
        row_lst = ['B' ,'C' , 'D' , 'E' , 'F' , 'G' , 'H' , 'I' ]

        # Changing the cell dimension and increasing the width to 15 and 20
        section_sheet.column_dimensions[f'A'].width = 15
        for i in range(0,8):
            # set the width of the column 
            section_sheet.column_dimensions[f'{row_lst[i]}'].width = 20


        # Write section timetable data
        row = 1
        for section_id, section_data in self.sections_timetable.items():

            # print section id first
            self.change_font(section_sheet,f"A{row}" , True)
            section_sheet[f'A{row}'] = section_id

            # Printing periods on sheet
            for i in range(0,8):
                self.change_font(section_sheet,f'{row_lst[i]}{row}')
                section_sheet[f'{row_lst[i]}{row}'] = i+1


            for day, day_data in section_data.items():

                row+=1
                # printing the Day name in the column B
                self.change_font(section_sheet,f"A{row}")
                section_sheet[f'A{row}'] = day
                
                row_lst_index = 0
                for period, class_data in day_data.items():

                    if class_data is not None:
                        # subject_code, teacher_ids, room_number = class_data
                        class_str=f"{class_data[0]} , {class_data[1]} , {class_data[2]}"
                        section_sheet[f'{row_lst[row_lst_index]}{row}'] = class_str
                    else:
                        section_sheet[f'{row_lst[row_lst_index]}{row}'] = "NA"
                    
                    self.change_font(section_sheet,f'{row_lst[row_lst_index]}{row}',isnormalvalue=True)

                    row_lst_index+=1
            row+=2
        workbook.save(self.file_path)

    def print_output_excel_teacher(self):
        
        self.clear_sheet("Teacher Timetable")
        workbook = openpyxl.load_workbook(self.file_path)
        teacher_sheet = workbook['Teacher Timetable']
        row_lst = ['B' ,'C' , 'D' , 'E' , 'F' , 'G' , 'H' , 'I' ]

        # Changing the cell dimension and increasing the width to 15 and 20
        teacher_sheet.column_dimensions[f'A'].width = 15
        for i in range(0,8):
            # set the width of the column 
            teacher_sheet.column_dimensions[f'{row_lst[i]}'].width = 20


        # Write section timetable data
        row = 1
        for teacher_id, teacher_data in self.teachers_timetable.items():

            # print section id first
            self.change_font(teacher_sheet,f"A{row}" , True)
            teacher_sheet[f'A{row}'] = teacher_id

            # Printing periods on sheet
            for i in range(0,8):
                self.change_font(teacher_sheet,f'{row_lst[i]}{row}')
                teacher_sheet[f'{row_lst[i]}{row}'] = i+1


            for day, day_data in teacher_data.items():

                row+=1
                # printing the Day name in the column B
                self.change_font(teacher_sheet,f"A{row}")
                teacher_sheet[f'A{row}'] = day
                
                row_lst_index = 0
                for period, class_data in day_data.items():

                    if class_data is not None:
                        # subject_code, teacher_ids, room_number = class_data
                        class_str=f"{class_data[0]} , {class_data[1]} , {class_data[2]}"
                        teacher_sheet[f'{row_lst[row_lst_index]}{row}'] = class_str
                    else:
                        teacher_sheet[f'{row_lst[row_lst_index]}{row}'] = "NA"
                    
                    self.change_font(teacher_sheet,f'{row_lst[row_lst_index]}{row}',isnormalvalue=True)

                    row_lst_index+=1
            row+=2
        workbook.save(self.file_path)

    def print_output_excel_room(self):
        
        self.clear_sheet("Room Timetable")
        workbook = openpyxl.load_workbook(self.file_path)
        room_sheet = workbook['Room Timetable']
        row_lst = ['B' ,'C' , 'D' , 'E' , 'F' , 'G' , 'H' , 'I' ]

        # Changing the cell dimension and increasing the width to 15 and 20
        room_sheet.column_dimensions[f'A'].width = 15
        for i in range(0,8):
            # set the width of the column 
            room_sheet.column_dimensions[f'{row_lst[i]}'].width = 20


        # Write section timetable data
        row = 1
        for room_no, room_data in self.rooms_timetable.items():

            # print section id first
            self.change_font(room_sheet,f"A{row}" , True)
            room_sheet[f'A{row}'] = room_no

            # Printing periods on sheet
            for i in range(0,8):
                self.change_font(room_sheet,f'{row_lst[i]}{row}')
                room_sheet[f'{row_lst[i]}{row}'] = i+1


            for day, day_data in room_data.items():

                row+=1
                # printing the Day name in the column B
                self.change_font(room_sheet,f"A{row}")
                room_sheet[f'A{row}'] = day
                
                row_lst_index = 0
                for period, class_data in day_data.items():

                    if class_data is not None:
                        # subject_code, teacher_ids, room_number = class_data
                        class_str=f"{class_data[0]} , {class_data[1]} , {class_data[2]}"
                        room_sheet[f'{row_lst[row_lst_index]}{row}'] = class_str
                    else:
                        room_sheet[f'{row_lst[row_lst_index]}{row}'] = "NA"
                    
                    self.change_font(room_sheet,f'{row_lst[row_lst_index]}{row}',isnormalvalue=True)

                    row_lst_index+=1
            row+=2
        workbook.save(self.file_path)


def main():
    excel_input_file = "Timetable.xlsx"

    # Data type needed for subjects - String,String,int,bool,bool,int,list
    # Data type needed for Teachers- String,String,list,list,bool,String,
    # Data type needed for Sections - String,String,list,int,int,string,string
    # Data type needed for Rooms - String,Bool,int

    # Take input from Excel
    timetable_data = read_data(excel_input_file)
    teachers,subjects, sections,rooms = timetable_data.read_all_data()

    # Create a timetable object
    timetable = Timetable(teachers,subjects, sections,rooms)

    # Generate the timetable
    timetable.generate_timetable()
    timetable.verify_timetable()

    get_output = excel_output(timetable,excel_input_file)
    get_output.print_output_excel_section()
    get_output.print_output_excel_teacher()
    get_output.print_output_excel_room()

    # Printing sections Timetable
    # for sec in sections:
    #     for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']:
    #         print("\nYear - " ,sec.year , "Section - " , sec.section_id , " and Day - ",day,"\n")
    #         pprint(timetable.section_timetable[sec.section_id][day])

    # # Printing Teachers Timetable
    # for tec in teachers:
    #     for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']:
    #         print("\nTeacher Name - " , tec.name , " and Day - ",day,"\n")
    #         pprint(timetable.teacher_timetable[tec.teacher_id][day])

    print("Total classes Expected for - 297 and Total classes Occuring - " , timetable.total_classes)

if __name__ == "__main__":
    main()
